# -------------------------ABOUT --------------------------

# pyinstaller --onefile app.py
# Tool: NCOP DB Import Manager Tool
# Developer: dyoliya
# Created: 2026-03-03

# © 2026 dyoliya. All rights reserved.

# ---------------------------------------------------------

import os
import threading
from tkinter import messagebox
from zoneinfo import ZoneInfo
import re
import pandas as pd
import shutil
import customtkinter as ctk
from openpyxl import load_workbook

from customtkinter import filedialog

from ncop_transform import clean_and_prepare_df
from ncop_sqlite import (
    connect_sqlite,
    ensure_table_and_columns,
    get_existing_columns,
    insert_rows,
)
from datetime import datetime
from datetime import date
from fractions import Fraction
from pathlib import Path
from typing import Optional, Tuple

CENTRAL_TZ = ZoneInfo("America/Chicago")

def get_daily_db_path(tool_prefix: str = "ncop") -> str:
    """
    Daily-versioned + cumulative, but detection is based ONLY on database/ folder.

    - Active DB lives in database/: YYYY-MM-DD-<tool_prefix>.db
    - If today's DB doesn't exist:
        - Find latest YYYY-MM-DD-<tool_prefix>.db in database/ ONLY
        - Copy it to today's DB (so today includes history)
        - Move all other .db files in database/ to database/previous_versions/
    - Does NOT look at previous_versions/ when choosing the base DB.
    """
    base_dir = Path(__file__).resolve().parent
    db_dir = base_dir / "database"
    prev_dir = db_dir / "previous_versions"
    db_dir.mkdir(parents=True, exist_ok=True)
    prev_dir.mkdir(parents=True, exist_ok=True)

    today_str = datetime.now(CENTRAL_TZ).strftime("%Y-%m-%d")
    today_name = f"{today_str}-{tool_prefix}.db"
    today_path = db_dir / today_name

    # Case 1: today's DB already exists
    if today_path.exists():
        for p in db_dir.glob("*.db"):
            if p.is_file() and p.name != today_name:
                p.rename(prev_dir / p.name)
        return str(today_path), None, "existing"

    # Find latest dated DB in database/
    pat = re.compile(rf"^(\d{{4}}-\d{{2}}-\d{{2}})-{re.escape(tool_prefix)}\.db$")
    candidates = []
    for p in db_dir.glob("*.db"):
        m = pat.match(p.name)
        if m:
            candidates.append((m.group(1), p))

    latest_path = sorted(candidates, key=lambda x: x[0])[-1][1] if candidates else None

    # Case 2: copy forward from latest base DB
    if latest_path and latest_path.exists():
        shutil.copy2(latest_path, today_path)
        copied_from = latest_path.name

        # archive other DBs
        for p in db_dir.glob("*.db"):
            if p.is_file() and p.name != today_name:
                p.rename(prev_dir / p.name)

        return str(today_path), copied_from, "copied"

    # Case 3: truly new (no base DB)
    # (No need to create the file here; connect_sqlite will create it when connecting.)
    for p in db_dir.glob("*.db"):
        if p.is_file() and p.name != today_name:
            p.rename(prev_dir / p.name)

    return str(today_path), None, "new"

def is_excel_hash_overflow(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    return bool(re.fullmatch(r"#+", s))  # "#######"

def table_exists(conn, table_name: str) -> bool:
    cur = conn.cursor()
    cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,)
    )
    return cur.fetchone() is not None


def get_expected_columns(conn, table_name: str) -> set[str]:
    # existing DB columns
    cols = set(get_existing_columns(conn, table_name))
    # ignore system columns you add automatically
    ignore = {"ncop_id", "date_uploaded"}
    return {c for c in cols if c not in ignore}


def validate_strict_schema(
    incoming_cols: list[str],
    expected_cols: set[str],
    original_cols: list[str] | None = None,
    sanitized_cols: list[str] | None = None,
):
    incoming_set = set(incoming_cols)

    missing = sorted(expected_cols - incoming_set)
    extra = sorted(incoming_set - expected_cols)

    if not missing and not extra:
        return  # ok

    # optional: show mapping help
    mapping_info = ""
    if original_cols is not None and sanitized_cols is not None:
        pairs = list(zip(original_cols, sanitized_cols))
        mapping_info = "\n\n[DEBUG] Column mapping (original → sanitized):\n" + "\n".join(
            f"  - {o}  →  {s}" for o, s in pairs[:80]
        )
        if len(pairs) > 80:
            mapping_info += "\n  ... (truncated)"

    msg_lines = ["Schema mismatch: incoming file columns do not match existing DB schema."]

    if missing:
        msg_lines.append("\nMissing required columns (expected in DB, not found in file):")
        msg_lines += [f"  - {c}" for c in missing]

    if extra:
        msg_lines.append("\nUnexpected columns (present in file, not in DB schema):")
        msg_lines += [f"  - {c}" for c in extra]

    raise ValueError("\n".join(msg_lines) + mapping_info)

def read_input_file(path: str, sheet_name: str | None = None) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()

    if ext == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False)

    if ext not in {".xlsx", ".xlsm", ".xls"}:
        raise ValueError("Unsupported file type. Please select a .csv or .xlsx file.")

    # --- Read via pandas first (fast) ---
    xls = pd.ExcelFile(path)
    sheet = sheet_name or xls.sheet_names[0]
    df = pd.read_excel(
        xls,
        sheet_name=sheet,
        dtype=str,
        keep_default_na=False,  # ✅ do not convert stuff into NaN
        na_filter=False,        # ✅ don’t treat blanks/tokens as NaN
    ).fillna("")
    
    PHONE_COLS = {
        "Phone1", "Phone2", "Phone3", "Phone4",
        "CLEANED PHONE1", "CLEANED PHONE2", "CLEANED PHONE3", "CLEANED PHONE4",
        "Phone 1", "Phone 2", "Phone 3", "Phone 4", "Phone 5",
    }

    HEIRSHIP_COL = "Heirship Report Link"
    # accept either spelling, case-insensitive
    OWNER_COL_CANDIDATES = {"OWNERHSIP PORTION", "OWNERSHIP PORTION"}

    def _norm_header(x) -> str:
        return str(x).strip().casefold() if x is not None else ""

    # Decide if we even need openpyxl
    needs_wb = (
        any(c in df.columns for c in PHONE_COLS)
        or (HEIRSHIP_COL in df.columns)
        or any(c in df.columns for c in OWNER_COL_CANDIDATES)
    )

    if not needs_wb:
        return df

    wb = None
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet]

        # Build header -> column index map (case-insensitive)
        header_to_col = {}
        for i, cell in enumerate(ws[1], start=1):
            header_to_col[_norm_header(cell.value)] = i

        n = len(df)

        # --- 1) Fix "#######" for phone columns only ---
        for col in PHONE_COLS:
            if col not in df.columns:
                continue

            excel_col_idx = header_to_col.get(_norm_header(col))
            if not excel_col_idx:
                continue

            for row_i in range(n):
                v = df.at[row_i, col]

                # pandas might give "", NaN, or "#######" depending on how it read the file
                is_blank_or_nan = (v == "") or pd.isna(v)
                is_hash = is_excel_hash_overflow(v)

                if is_blank_or_nan or is_hash:
                    xl_cell = ws.cell(row=2 + row_i, column=excel_col_idx)
                    raw = xl_cell.value

                    # store raw cell value as text
                    if raw is None:
                        df.at[row_i, col] = ""
                    elif isinstance(raw, (int, float)):
                        # avoid "9723612930.0"
                        df.at[row_i, col] = str(int(raw)) if float(raw).is_integer() else str(raw)
                    else:
                        df.at[row_i, col] = str(raw)

        # --- 2) Extract real hyperlink targets for Heirship Report Link ---
        if HEIRSHIP_COL in df.columns:
            excel_col_idx = header_to_col.get(_norm_header(HEIRSHIP_COL))
            if excel_col_idx:
                real_links = []
                for r in range(2, 2 + n):
                    cell = ws.cell(row=r, column=excel_col_idx)
                    if cell.hyperlink and cell.hyperlink.target:
                        real_links.append(cell.hyperlink.target)
                    else:
                        val = cell.value
                        real_links.append("" if val is None else str(val))
                df[HEIRSHIP_COL] = real_links

        # --- 3) Fix Ownership Portion fractions being misread as dates (1/4, 1/2) ---
        # --- 3) Fix Ownership Portion fractions being misread as dates (1/4, 1/2, 1/3) ---
        owner_col_in_df = None
        for candidate in OWNER_COL_CANDIDATES:
            if candidate in df.columns:
                owner_col_in_df = candidate
                break

        def ownership_cell_to_text(cell) -> str:
            v = cell.value
            if v is None:
                return ""

            # If Excel stored it as a DATE, convert back to M/D (Jan 3 => "1/3")
            if getattr(cell, "is_date", False) and isinstance(v, (datetime, date)):
                return f"{v.month}/{v.day}"

            # Already text (best case)
            if isinstance(v, str):
                return v.strip()

            # Numeric: try to express as fraction if it looks like a portion
            if isinstance(v, (int, float)):
                # Keep percent as percent if it was entered that way
                if "%" in str(getattr(cell, "number_format", "")):
                    return f"{v * 100:.0f}%"

                # If it's between 0 and 1, it's likely a portion; turn into a nice fraction
                if 0 < float(v) < 1:
                    frac = Fraction(v).limit_denominator(256)
                    return f"{frac.numerator}/{frac.denominator}"

                return str(v)

            return str(v)

        if owner_col_in_df:
            excel_col_idx = header_to_col.get(_norm_header(owner_col_in_df))
            if excel_col_idx:
                fixed_vals = []
                for r in range(2, 2 + n):
                    cell = ws.cell(row=r, column=excel_col_idx)
                    fixed_vals.append(ownership_cell_to_text(cell))
                df[owner_col_in_df] = fixed_vals

        return df  # ✅ IMPORTANT: return the DataFrame
    
    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass

class NCOPImporterApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # --- Signature UI colors (your theme) ---
        self.PANEL_BG = "#273946"
        self.APP_BG = "#fff6de"
        self.ACCENT = "#CB1F47"
        self.ACCENT_HOVER = "#ffab4c"
        self.TEXT_DARK = "#273946"

        self.title("NCOP: DB Import Manager [demo_v1]")
        self.geometry("430x720")
        self.resizable(False, True)
        self.minsize(430, 650)

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")
        self.configure(fg_color=self.APP_BG)

        self.main_frame = ctk.CTkFrame(self, fg_color=self.APP_BG, corner_radius=12)
        self.main_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.title_label = ctk.CTkLabel(
            self.main_frame,
            text="NCOP DB Import Manager Tool",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=self.TEXT_DARK
        )
        self.title_label.pack(pady=(12, 6))

        input_tab = self._create_locked_tab_section(title="I m p o r t", height=200)
        self._setup_import_tab(input_tab)

        self.progress = ctk.CTkProgressBar(
            self.main_frame,
            width=390,
            fg_color=self.PANEL_BG,
            progress_color=self.ACCENT
        )
        self.progress.set(0)
        self.progress.pack(pady=10)

        self.log_container = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.log_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.log_container.grid_rowconfigure(1, weight=1)
        self.log_container.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.log_container,
            text="Activity Log",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=self.TEXT_DARK
        ).grid(row=0, column=0, sticky="w", padx=10, pady=(0, 4))

        self.log_box = ctk.CTkTextbox(self.log_container, fg_color="#ffffff", text_color=self.TEXT_DARK)
        self.log_box.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
        self.log_box.configure(state="disabled")

        self.input_paths = []

                # github link (hihi)
        self.credit_label = ctk.CTkLabel(
        self,
        text="© dyoliya • GitHub",
        text_color="#484949",
        font=ctk.CTkFont(size=8, underline=False),
        cursor="hand2"
        )
        self.credit_label.place(relx=1.0, x=-10, y=1, anchor="ne") 
        self.credit_label.bind("<Button-1>", lambda e: self.open_url("https://github.com/dyoliya"))


    def _create_locked_tab_section(self, title: str, height: int):
        tab_font = ctk.CTkFont(size=12, weight="bold")
        tv = ctk.CTkTabview(self.main_frame, width=390, height=height)
        tv.configure(
            fg_color=self.PANEL_BG,
            segmented_button_fg_color=self.APP_BG,
            segmented_button_selected_color=self.PANEL_BG,
            segmented_button_selected_hover_color=self.PANEL_BG,
            segmented_button_unselected_color=self.PANEL_BG,
            text_color=self.ACCENT_HOVER,
            text_color_disabled=self.ACCENT_HOVER
        )
        tv.pack(fill="x", padx=10, pady=(10, 8), anchor="w")
        tv.configure(anchor="w")

        tab = tv.add(title)

        try:
            tv._segmented_button.grid_configure(sticky="w")
            btn = tv._segmented_button._buttons_dict[title]
            btn.configure(width=140, height=35)
            tv._segmented_button.configure(state="disabled", font=tab_font)
            for b in tv._segmented_button._buttons_dict.values():
                b.configure(state="disabled")
        except Exception:
            pass

        return tab

    def _setup_import_tab(self, tab):
        LABEL_W = 130

        # Row 1: Browse
        row1 = ctk.CTkFrame(tab, fg_color="transparent")
        row1.pack(fill="x", padx=10, pady=(10, 6), anchor="w")

        ctk.CTkLabel(
            row1,
            text="Select Files:",
            width=LABEL_W,
            anchor="w",
            text_color=self.APP_BG
        ).pack(side="left")

        self.file_label = ctk.CTkLabel(row1, text="(none)", anchor="w", text_color=self.APP_BG)
        self.file_label.pack(side="left", padx=(0, 8), fill="x", expand=True)

        self.pick_file_btn = ctk.CTkButton(
            row1,
            text="Browse",
            width=80,
            fg_color=self.ACCENT,
            hover_color=self.ACCENT_HOVER,
            command=self.pick_input_file
        )
        self.pick_file_btn.pack(side="right")

        # ✅ Selected files panel (dark + scroll)
        panel = ctk.CTkFrame(tab, fg_color="#1e2b34", corner_radius=10)
        # panel.pack(fill="both", expand=False, padx=10, pady=(6, 8))
        panel.pack(fill="x", expand=False, padx=10, pady=(6, 8))

        panel_header = ctk.CTkFrame(panel, fg_color="transparent")
        panel_header.pack(fill="x", padx=10, pady=(8, 4))

        ctk.CTkLabel(
            panel_header,
            text="Selected files",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color="#ffffff"
        ).pack(side="left")

        self.clear_btn = ctk.CTkButton(
            panel_header,
            text="Clear",
            width=70,
            fg_color=self.PANEL_BG,
            hover_color="#334957",
            command=self.clear_selected_files
        )
        self.clear_btn.pack(side="right")

        # scrollable list WRAPPER (controls the height)
        list_wrap = ctk.CTkFrame(panel, fg_color="#1e2b34", height=80, corner_radius=10)
        list_wrap.pack(fill="x", expand=False, padx=10, pady=(0, 10))
        list_wrap.pack_propagate(False)  # force the wrapper height

        self.files_list = ctk.CTkScrollableFrame(
            list_wrap,
            fg_color="#1e2b34",
            scrollbar_fg_color="#273946",
            scrollbar_button_color=self.ACCENT,
            scrollbar_button_hover_color=self.ACCENT_HOVER,
        )
        self.files_list.pack(fill="both", expand=True)

        # Import button
        self.import_btn = ctk.CTkButton(
            tab,
            text="Import & Append to DB",
            fg_color=self.ACCENT,
            hover_color=self.ACCENT_HOVER,
            command=self.start_import
        )
        self.import_btn.pack(pady=(6, 12), padx=10)   # 12px space below

    def clear_selected_files(self):
        self.input_paths = []
        self.file_label.configure(text="(none)")
        self._refresh_files_list()
        self._log("[INFO] Cleared selected files.\n")

    def _refresh_files_list(self):
        # remove existing rows
        for child in self.files_list.winfo_children():
            child.destroy()

        if not self.input_paths:
            ctk.CTkLabel(
                self.files_list,
                text="No files selected",
                text_color="#cbd5e1"
            ).pack(anchor="w", pady=6)
            return

        for p in self.input_paths:
            row = ctk.CTkFrame(self.files_list, fg_color="transparent")
            row.pack(fill="x", pady=3)

            ctk.CTkLabel(
                row,
                text=os.path.basename(p),
                text_color="#ffffff",
                anchor="w"
            ).pack(side="left", fill="x", expand=True)

            ctk.CTkLabel(
                row,
                text=os.path.dirname(p),
                text_color="#94a3b8",
                anchor="e"
            ).pack(side="right")

    # ---------- log/progress helpers ----------
    def _log(self, text: str):
        def _append():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", text + "\n")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _append)

    def _divider(self):
        self._log("- - - - - - - - - - - - - - - - - - - - - - - - - - -")

    def progress_callback(self, fraction, msg=None):
        self.progress.set(max(0.0, min(1.0, float(fraction))))
        if msg:
            self._log(msg)
        self.update_idletasks()

    def _ui_error(self, title: str, msg: str):
        self.after(0, lambda m=msg: messagebox.showerror(title, m))

    # ---------- UI actions ----------
    def pick_input_file(self):
        paths = filedialog.askopenfilenames(
            title="Select input file(s)",
            filetypes=[("CSV or Excel", "*.csv *.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        if not paths:
            return

        self.input_paths = list(paths)
        self.file_label.configure(text=f"{len(self.input_paths)} file(s) selected")
        self._refresh_files_list()  # ✅ update list UI

        self._log("[INFO] Selected inputs:\n  - " + "\n  - ".join(self.input_paths))
        self._log("\n")

    def start_import(self):
        self._divider()

        if not self.input_paths:
            messagebox.showwarning("Missing input", "Please select one or more CSV/XLSX files.")
            return

        db_path, copied_from, mode = get_daily_db_path("ncop")
        self._log(f"[DB] Using daily DB: {db_path}\n")

        if mode == "copied":
            self._log(f"[DB] Copied base DB: {copied_from}\n")
        elif mode == "existing":
            self._log("[DB] Today's DB already exists. Continuing append..\n")
        else:  # "new"
            self._log("[DB] No base DB found. Starting a NEW DB for today.\n")

        table = "ncop"

        self.import_btn.configure(state="disabled")
        self.pick_file_btn.configure(state="disabled")
        self.progress_callback(0, "Starting import...\n")

        threading.Thread(
            target=self._import_worker,
            args=(db_path, self.input_paths, table),
            daemon=True
        ).start()

    def _import_worker(self, db_path, input_paths, table):

        try:
            conn = connect_sqlite(db_path)
            try:
                total_files = len(input_paths)
                audit_chunks = []  # collects audit_df from each input file

                for idx, input_path in enumerate(input_paths, start=1):
                    self.progress_callback(0, f"[{idx}/{total_files}] Reading: {os.path.basename(input_path)}\n")
                    df_raw = read_input_file(input_path)  # ✅ no sheet_name
                    
                    if df_raw.empty:
                        self._log(f"[SKIP] {os.path.basename(input_path)} has no rows.\n")
                        continue

                    self.progress_callback(0, f"[{idx}/{total_files}] Cleaning/standardizing...\n")
                    df, original_cols, sanitized_cols, phone_cols, date_cols, audit_df = clean_and_prepare_df(df_raw)

                    if audit_df is not None and not audit_df.empty:
                        audit_df = audit_df.copy()
                        audit_df.insert(0, "source_file", os.path.basename(input_path))
                        audit_chunks.append(audit_df)
                        
                    if not table_exists(conn, table):
                        # first-ever import: create table + columns
                        self._log("[SCHEMA] Table not found. Creating schema from this first file...\n")
                        ensure_table_and_columns(conn, table, list(df.columns))
                    else:
                        # strict mode: validate only (no new columns allowed)
                        expected = get_expected_columns(conn, table)
                        validate_strict_schema(
                            incoming_cols=list(df.columns),
                            expected_cols=expected,
                            original_cols=original_cols,
                            sanitized_cols=sanitized_cols,
                        )
                        self._log("[SCHEMA] OK: incoming columns match existing DB schema.\n")

                    self._log(f"[{idx}/{total_files}] {os.path.basename(input_path)}: {len(df):,} rows\n")
                    self._log(f"[MAP] Columns: {len(original_cols)} original → {len(sanitized_cols)} sanitized\n")

                    # if phone_cols:
                    #     self._log(f"[TYPE] Phone columns: {', '.join(phone_cols[:12])}" + ("..." if len(phone_cols) > 12 else ""))
                    # if date_cols:
                    #     self._log(f"[DATE] Date columns: {', '.join(date_cols[:12])}" + ("..." if len(date_cols) > 12 else ""))

                    self.progress_callback(0, f"[{idx}/{total_files}] Preparing table...\n")
                    ensure_table_and_columns(conn, table, list(df.columns))

                    existing_cols = get_existing_columns(conn, table)
                    if "ncop_id" not in existing_cols:
                        self._log("[WARN] Existing table has no ncop_id PK. Inserts will still work, but PK will not be auto-added.\n")
                    if "date_uploaded" not in existing_cols:
                        self._log("[WARN] date_uploaded was missing and should have been added.\n")

                    self.progress_callback(0, f"[{idx}/{total_files}] Inserting rows...\n")
                    insert_rows(
                        conn,
                        table,
                        df,
                        progress_callback=lambda frac: self.progress_callback(frac),
                        tz=CENTRAL_TZ
                    )

                # --- Save audit report (if any) ---
                if audit_chunks:
                    audit_all = pd.concat(audit_chunks, ignore_index=True)

                    out_dir = Path(__file__).resolve().parent / "output"
                    out_dir.mkdir(parents=True, exist_ok=True)

                    ts = datetime.now(CENTRAL_TZ).strftime("%Y-%m-%d_%H%M%S")
                    audit_path = out_dir / f"ncop_null_audit_{ts}.xlsx"

                    with pd.ExcelWriter(audit_path, engine="openpyxl") as writer:
                        audit_all.to_excel(writer, sheet_name="null_conversions", index=False)

                    self._log(f"[AUDIT] Saved NULL conversion audit: {audit_path}\n")
                else:
                    self._log("[AUDIT] No suspicious NULL conversions detected.\n")
                    
                self._log(f"[DONE] Daily DB updated: {os.path.basename(db_path)} → table '{table}'\n")
                self.after(0, lambda: messagebox.showinfo("Success", f"Imported into:\n{db_path}\n\nTable: {table}"))

            finally:
                conn.close()

        except Exception as e:
            err_msg = str(e)
            self._log(f"[ERROR] {err_msg}\n")
            self._ui_error("Error", err_msg)

        finally:
            self.after(0, lambda: self.import_btn.configure(state="normal"))
            self.after(0, lambda: self.pick_file_btn.configure(state="normal"))
            self._divider()
            self.progress_callback(0, "Waiting for action...")


if __name__ == "__main__":
    app = NCOPImporterApp()
    app.mainloop()